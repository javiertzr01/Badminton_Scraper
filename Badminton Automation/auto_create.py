import numpy as np
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Color, PatternFill, Font
from openpyxl.styles.colors import Color
from pandas.core import series

file_name = "SUTD Badminton Recre Session (15th Nov)(1-49).xlsx"    ###CHANGE THIS EVERY TIME###
declaration_name = "OSL_Badminton_Declaration_2021T3wef12Oct_Date.xlsx"
final_declaration_name = "OSL_Badminton_Declaration_2021T3wef12Oct_Date - Copy.xlsx"
df = pd.read_excel(rf"C:\Badminton_Auto\{file_name}", usecols="N,Q,T,W,Z,AC,AF,AI,AL,AO,AR,AU,AX,BA,BD,BJ,BM,BP")
df_declaration = pd.read_excel(rf"C:\Badminton_Auto\{declaration_name}", header=[0,1])
wb = load_workbook(rf"C:\Badminton_Auto\{final_declaration_name}")
ws = wb.active
# term1_columns = ["Q","T","W","Z","AC","BP"]
# term3_columns = ["AF","AI","AL","AO","AR","BP"]
# others_columns = ["AU","AX","BA","BD","BJ","BM","BP"]
session1_size = 0
session2_size = 0
vr = open("Vaccination_Reports.txt", "w")

pd.set_option('display.max_columns', 10)

columns = pd.DataFrame(df_declaration.columns.tolist())
columns.loc[columns[1].str.startswith('Unnamed:'), 1] = np.nan
columns = columns.fillna(method='ffill', axis=1)
df_declaration.columns = pd.MultiIndex.from_tuples(columns.to_records(index=False).tolist())


#print(df_declaration)

#0 = Name
#1 = ID
#2 = Hostel
#3 = Time Slot
#4 = First Time
#5 = Submitted Vaccination Report

### Trying to add rows###
# for n in range(df_declaration.shape[0] + 1, df.shape[0] + 1):
#     df_add = df_declaration.iloc[20,:]
#     df_add["S/no"] = n
#     df_declaration.append(df_add,ignore_index=True)

for n in range(df.shape[0]):
    info = df.loc[n,:]
    if info["Time Slot?"] == "7pm - 8:50pm" or info["Time Slot?2"] == "7pm - 8:50pm" or info["Time Slot?3"] == "7pm - 8:50pm":
        if session1_size < 20:
            index = session1_size
            session1_size += 1
        else:
            continue
    elif info["Time Slot?"] == "9:10pm -11pm" or info["Time Slot?2"] == "9:10pm -11pm" or info["Time Slot?3"] == "9:10pm -11pm":
        if session2_size < 20:
            index = session2_size + 20
            session2_size += 1
        else:
            continue
    d_name = df_declaration.loc[index, ("Full name", "Full name")]
    d_student_id = df_declaration.loc[index, ("Student ID", "Student ID")]
    d_hostel_resident = df_declaration.loc[index, ("Type of student (please check)", "Hostel Resident")]
    d_term1 = df_declaration.loc[index, ("Type of student (please check)", "Term 1 ")]
    d_term3 = df_declaration.loc[index, ("Type of student (please check)", "Term 3 ")]
    d_pr = df_declaration.loc[index, ("Type of student (please check)", "Postgraduate Research")]
    d_masters = df_declaration.loc[index, ("Type of student (please check)", "Postgraduate Coursework & Masters Students\n(After 6pm from Weeks 1 to 6 only) ")]
    d_others = df_declaration.loc[index, ("Type of student (please check)", "Other Face-to-Face Classes \n\nPls indicate Pillar, Term & Subject Code")]
    
    
    
    if info["Are you a freshmore?"] == "Term 1 Freshmore":
        name = info["Full Name:"]
        student_id = str(int(info["Student ID"]))
        hostel_resident = info["Hostel Resident?"]
        first_time = info["First Time Joining a Session in this term?"]
        df_declaration.loc[index, ("Type of student (please check)", "Term 1 ")] = 1

    elif df.loc[n,:]["Are you a freshmore?"] == "Term 3 Freshmore":
        name = info["Full Name:2"]
        student_id = str(int(info["Student ID2"]))
        hostel_resident = info["Hostel Resident?2"]
        first_time = info["First time joining a session in this term?2"]
        df_declaration.loc[index, ("Type of student (please check)", "Term 3 ")] = 1
        
    elif df.loc[n,:]["Are you a freshmore?"] == "No, I am Postgraduate Research/Postgraduate Coursework & Masters/Others":
        name = info["Full Name:3"]
        student_id = str(int(info["Student ID3"]))
        category = info["What category are you in SUTD?"]
        hostel_resident = info["Hostel Resident?3"]
        first_time = info["First time joining a session in this term?3"]
        if category == "Postgraduate Research":
            df_declaration.loc[index, ("Type of student (please check)", "Postgraduate Research")] = 1
        elif category == "Postgraduate Coursework & Masters Student":
            df_declaration.loc[index, ("Type of student (please check)", "Postgraduate Coursework & Masters Students\n(After 6pm from Weeks 1 to 6 only) ")] = 1
        elif category == "Other Face-to-Face Classes (Please Elaborate Below)":
            df_declaration.loc[index, ("Type of student (please check)", "Other Face-to-Face Classes \n\nPls indicate Pillar, Term & Subject Code")] = 1
    
    if first_time == "Yes":
        submitted = info["Have you submitted your REDACTED Vaccination Report?"]
        if submitted == "Yes":
            vr.write(name + "    " + "DONE\n")
        elif submitted == "No":
            vr.write(name + "    " + "NOT YET\n")
    
    df_declaration.loc[index, ("Full name", "Full name")] = name
    df_declaration.loc[index, ("Student ID", "Student ID")] = student_id
    df_declaration.loc[index, ("Type of student (please check)", "Hostel Resident")] = 1 if hostel_resident == "Yes" else None 
    df_declaration.loc[index, ("Vaccination Status for Indoor Activities \n(indicate A, B, or C and submit relevant doc)", "Vaccination Status for Indoor Activities \n(indicate A, B, or C and submit relevant doc)")] = "A"


#Input information
rows = dataframe_to_rows(df_declaration, index=False, header=False)
for r_idx, row in enumerate(rows, 3):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        if (cell.value == 1 and c_idx != 1) or c_idx == 3 or c_idx == 4:
            cell.alignment = Alignment(horizontal='center',vertical='center')
            if cell.value == 1:
                cell.fill = PatternFill(start_color= "C6EFCE", end_color= "C6EFCE", fill_type="solid")
                cell.font = Font(color = "006100")
wb.save(final_declaration_name)